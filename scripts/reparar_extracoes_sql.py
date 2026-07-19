from __future__ import annotations

import argparse
import base64
import csv
import json
import os
import subprocess
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DATABASE_URL = "postgresql://postgres@127.0.0.1:5433/dark_jutsu"


def find_pg_bin() -> Path:
    candidates = [
        os.environ.get("PG_BIN"),
        r"C:\DarkJutsu\PostgreSQL\pgsql\bin",
        str(Path.home() / "Desktop" / "aplicacoes code" / "pgsql" / "bin"),
        str(Path.home() / "Desktop" / "postgresql-18.4-2-windows-x64-binaries" / "pgsql" / "bin"),
    ]
    for candidate in candidates:
        if candidate and (Path(candidate) / "psql.exe").exists():
            return Path(candidate)
    raise RuntimeError("psql.exe nao encontrado. Defina PG_BIN com a pasta bin do PostgreSQL.")


def database_parts(url: str) -> dict[str, str]:
    from urllib.parse import unquote, urlparse

    parsed = urlparse(url)
    user = unquote(parsed.username or "postgres")
    password = unquote(parsed.password or "")
    return {
        "host": parsed.hostname or "127.0.0.1",
        "port": str(parsed.port or 5432),
        "user": user,
        "password": password,
        "database": (parsed.path or "/dark_jutsu").lstrip("/"),
    }


class Psql:
    def __init__(self, database_url: str) -> None:
        self.parts = database_parts(database_url)
        self.psql = find_pg_bin() / "psql.exe"

    def base_cmd(self) -> list[str]:
        return [
            str(self.psql),
            "-X",
            "-h",
            self.parts["host"],
            "-p",
            self.parts["port"],
            "-U",
            self.parts["user"],
            "-d",
            self.parts["database"],
            "-v",
            "ON_ERROR_STOP=1",
        ]

    def env(self) -> dict[str, str]:
        env = os.environ.copy()
        env["PGPASSWORD"] = self.parts["password"]
        return env

    def scalar_rows(self, sql: str, separator: str = "\t") -> list[str]:
        proc = subprocess.run(
            [*self.base_cmd(), "-At", "-F", separator, "-c", sql],
            text=True,
            capture_output=True,
            env=self.env(),
        )
        if proc.returncode:
            raise RuntimeError(proc.stderr.strip() or proc.stdout.strip())
        return [line for line in proc.stdout.splitlines() if line.strip()]

    def run_file(self, path: Path) -> None:
        proc = subprocess.run([*self.base_cmd(), "-f", str(path)], text=True, capture_output=True, env=self.env())
        if proc.returncode:
            raise RuntimeError(proc.stderr.strip() or proc.stdout.strip())


def psql_path(path: Path) -> str:
    return path.resolve().as_posix().replace("'", "''")


def json_text(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def num(value: Any) -> Any:
    if isinstance(value, bool) or value in ("", None):
        return None
    try:
        return float(value)
    except Exception:
        return None


def timestamp_iso(value: Any) -> str | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        raw = float(value)
        if raw > 100000000000:
            raw = raw / 1000
        return datetime.fromtimestamp(raw, timezone.utc).isoformat()
    return None


def decode_legacy_key(value: str) -> str:
    try:
        padded = value + "=" * (-len(value) % 4)
        decoded = base64.urlsafe_b64decode(padded.encode("ascii")).decode("utf-8")
        return decoded or value
    except Exception:
        return value


def write_tsv(path: Path, rows: list[list[Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t", lineterminator="\n")
        for row in rows:
            writer.writerow(["\\N" if value is None else value for value in row])


def workbook_limits(source: Path) -> list[list[Any]]:
    wb = load_workbook(source, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        header = [str(value).strip().lower() if value is not None else "" for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        index = {name: idx for idx, name in enumerate(header)}
        rows: list[list[Any]] = []
        for line in ws.iter_rows(min_row=2, values_only=True):
            protheus = clean_text(line[index["protheus"]]) if "protheus" in index and index["protheus"] < len(line) else None
            cooperat = clean_text(line[index["cooperat"]]) if "cooperat" in index and index["cooperat"] < len(line) else None
            if not protheus and not cooperat:
                continue
            minimo = num(line[index["minimo"]]) if "minimo" in index and index["minimo"] < len(line) else None
            maximo = num(line[index["maximo"]]) if "maximo" in index and index["maximo"] < len(line) else None
            reposicao_idx = index.get("reposicao", index.get("reposição"))
            reposicao = num(line[reposicao_idx]) if reposicao_idx is not None and reposicao_idx < len(line) else None
            saldo_anterior = num(line[index["saldo anterior"]]) if "saldo anterior" in index and index["saldo anterior"] < len(line) else None
            raw = {
                "protheus": protheus,
                "cooperat": cooperat,
                "saldoAnterior": saldo_anterior,
                "minimo": minimo,
                "maximo": maximo,
                "reposicao": reposicao,
                "fonte": str(source),
            }
            rows.append([protheus, cooperat, minimo, maximo, reposicao, saldo_anterior, json_text(raw)])
        return rows
    finally:
        wb.close()


def repair_cooperat(work_dir: Path, source: Path, psql: Psql) -> dict[str, Any]:
    raw = json.loads(source.read_text(encoding="utf-8"))
    codes = raw.get("codigos") if isinstance(raw.get("codigos"), dict) else {}
    import_run_id = str(uuid.uuid4())
    code_rows: list[list[Any]] = []
    event_rows: list[list[Any]] = []

    for code_key, item in codes.items():
        if not isinstance(item, dict):
            continue
        code = clean_text(item.get("codigo")) or str(code_key)
        code_raw = dict(item)
        code_raw.pop("eventos", None)
        code_rows.append(
            [
                code,
                clean_text(item.get("descricaoMaisRecente")),
                int(item.get("totalEventos") or 0),
                num(item.get("totalQuantidadeCompra")),
                num(item.get("totalQuantidadeSolicitada")),
                num(item.get("totalQuantidadeFornecida")),
                num(item.get("totalValorBaixa")),
                clean_text(item.get("primeiraData")),
                clean_text(item.get("ultimaData")),
                num(item.get("mediaQuantidadeCompra")),
                num(item.get("mediaQuantidadeSolicitada")),
                num(item.get("mediaQuantidadeFornecida")),
                num(item.get("mediaValorBaixa")),
                import_run_id,
                json_text(code_raw),
            ]
        )
        for event in item.get("eventos") if isinstance(item.get("eventos"), list) else []:
            if not isinstance(event, dict):
                continue
            event_rows.append(
                [
                    code,
                    clean_text(event.get("requisicao")),
                    clean_text(event.get("data")),
                    clean_text(event.get("dataBr")),
                    clean_text(event.get("descricao")),
                    clean_text(event.get("unidade")),
                    num(event.get("qtdSolicitada")),
                    num(event.get("qtdFornecida")),
                    num(event.get("valorBaixa")),
                    num(event.get("quantidadeCompra")),
                    clean_text(event.get("fonte")),
                    clean_text(event.get("origem")),
                    import_run_id,
                    json_text(event),
                ]
            )

    codes_file = work_dir / "cooperat_codes.tsv"
    events_file = work_dir / "cooperat_events.tsv"
    write_tsv(codes_file, code_rows)
    write_tsv(events_file, event_rows)

    metadata = {key: value for key, value in raw.items() if key != "codigos"}
    sql = f"""
BEGIN;
select set_config('app.role', 'service', true);
delete from cooperat_purchase_events;
delete from cooperat_purchase_codes;
insert into cooperat_import_runs (
  id, generated_at, description, quantity_rule, value_rule, event_limit_per_code,
  source_files, total_codes, total_events, source_hash, raw_data
) values (
  '{import_run_id}', nullif('{clean_text(raw.get("geradoEm")) or ""}', '')::timestamptz,
  {sql_literal(raw.get("descricao"))}, {sql_literal(raw.get("regraQuantidade"))},
  {sql_literal(raw.get("regraValor"))}, {int(raw.get("limiteEventosPorCodigo") or 0)},
  {sql_literal(json_text(raw.get("fontes") or []))}::jsonb,
  {int(raw.get("totalCodigos") or len(code_rows))}, {int(raw.get("totalEventos") or len(event_rows))},
  null, {sql_literal(json_text(metadata))}::jsonb
);
\\copy cooperat_purchase_codes (code, latest_description, total_events, total_purchase_qty, total_requested_qty, total_supplied_qty, total_low_value, first_date, last_date, avg_purchase_qty, avg_requested_qty, avg_supplied_qty, avg_low_value, import_run_id, raw_data) FROM '{psql_path(codes_file)}' WITH (FORMAT csv, DELIMITER E'\\t', NULL '\\N')
\\copy cooperat_purchase_events (code, requisition, event_date, event_date_label, description, unit, requested_qty, supplied_qty, low_value, purchase_qty, source, origin, import_run_id, raw_data) FROM '{psql_path(events_file)}' WITH (FORMAT csv, DELIMITER E'\\t', NULL '\\N')
COMMIT;
"""
    script = work_dir / "repair_cooperat.sql"
    script.write_text(sql, encoding="utf-8")
    psql.run_file(script)
    return {"codes": len(code_rows), "events": len(event_rows), "import_run_id": import_run_id}


def sql_literal(value: Any) -> str:
    if value is None:
        return "null"
    return "'" + str(value).replace("'", "''") + "'"


def load_recovered_history(psql: Psql) -> tuple[str, str, dict[str, Any]]:
    sql = """
select id::text, saved_at::text, payload->'historicoSaldo'
from inventory_snapshots s
where (
  select count(*) from jsonb_object_keys(coalesce(s.payload->'historicoSaldo', '{}'::jsonb))
) > 0
order by saved_at desc
limit 1
"""
    rows = psql.scalar_rows(sql)
    if not rows:
        return "", "", {}
    source_id, saved_at, payload_json = rows[0].split("\t", 2)
    return source_id, saved_at, json.loads(payload_json)


def repair_history(work_dir: Path, psql: Psql) -> dict[str, Any]:
    source_id, saved_at, history = load_recovered_history(psql)
    rows: list[list[Any]] = []
    for encoded_key, events in history.items():
        if not isinstance(events, list):
            continue
        item_key = decode_legacy_key(str(encoded_key))
        for event in events:
            if not isinstance(event, dict):
                continue
            rows.append(
                [
                    str(encoded_key),
                    item_key,
                    timestamp_iso(event.get("timestamp")),
                    clean_text(event.get("data")),
                    num(event.get("saldoAnterior")),
                    num(event.get("saldoAtual")),
                    num(event.get("delta")),
                    clean_text(event.get("tipo")),
                    json_text(event),
                ]
            )

    history_file = work_dir / "inventory_balance_history.tsv"
    write_tsv(history_file, rows)
    sql = f"""
BEGIN;
select set_config('app.role', 'service', true);
delete from inventory_balance_history where source = 'repair:inventory_snapshots/historicoSaldo';
create temp table tmp_repair_history (
  encoded_key text,
  item_key text,
  event_at timestamptz,
  event_date_label text,
  previous_balance numeric,
  current_balance numeric,
  delta numeric,
  event_type text,
  raw_data jsonb
) on commit drop;
\\copy tmp_repair_history FROM '{psql_path(history_file)}' WITH (FORMAT csv, DELIMITER E'\\t', NULL '\\N')
insert into inventory_balance_history (
  item_id, item_legacy_key, event_at, event_date_label, previous_balance,
  current_balance, delta, event_type, source, raw_data
)
select distinct on (h.encoded_key, h.item_key, h.event_at, h.delta, h.previous_balance, h.current_balance)
  i.id, h.item_key, h.event_at, h.event_date_label, h.previous_balance,
  h.current_balance, h.delta, h.event_type, 'repair:inventory_snapshots/historicoSaldo', h.raw_data
from tmp_repair_history h
left join inventory_items i
  on i.legacy_key = h.item_key
  or i.protheus_code = h.item_key
  or i.protheus_key = h.item_key
  or i.cooperat_code = h.item_key;
update inventory_snapshots latest
set payload = jsonb_set(latest.payload, '{{historicoSaldo}}', source.payload->'historicoSaldo', true),
    raw_metadata = jsonb_set(
      latest.raw_metadata,
      '{{historicoSaldoRecuperado}}',
      jsonb_build_object('source_snapshot_id', source.id::text, 'source_saved_at', source.saved_at::text, 'events', {len(rows)}),
      true
    )
from inventory_snapshots source
where latest.id = (select id from inventory_snapshots order by saved_at desc limit 1)
  and source.id = '{source_id}'::uuid
  and (
    select count(*) from jsonb_object_keys(coalesce(latest.payload->'historicoSaldo', '{{}}'::jsonb))
  ) = 0;
COMMIT;
"""
    script = work_dir / "repair_history.sql"
    script.write_text(sql, encoding="utf-8")
    psql.run_file(script)
    return {"source_snapshot_id": source_id, "source_saved_at": saved_at, "keys": len(history), "events": len(rows)}


def repair_limits(work_dir: Path, source: Path, psql: Psql) -> dict[str, Any]:
    workbook_rows = workbook_limits(source)
    workbook_file = work_dir / "inventory_item_limits_workbook.tsv"
    write_tsv(workbook_file, workbook_rows)
    sql = f"""
BEGIN;
select set_config('app.role', 'service', true);
delete from inventory_item_limits where source in ('repair:inventory_items.raw_data', 'repair:downloads/estoque_minimo.xlsx');
insert into inventory_item_limits (
  item_id, item_legacy_key, source, min_qty, max_qty, reorder_qty,
  previous_balance, applied, raw_data
)
select
  id,
  legacy_key,
  'repair:inventory_items.raw_data',
  nullif(raw_data->>'minimo', '')::numeric,
  nullif(raw_data->>'maximo', '')::numeric,
  nullif(raw_data->>'reposicao', '')::numeric,
  null,
  true,
  jsonb_build_object(
    'minimo', raw_data->'minimo',
    'maximo', raw_data->'maximo',
    'reposicao', raw_data->'reposicao',
    'origem', coalesce(raw_data->>'limitesOrigem', raw_data->>'minimoOrigem', 'snapshot')
  )
from inventory_items
where raw_data ? 'minimo'
  and raw_data ? 'maximo'
  and raw_data ? 'reposicao';
create temp table tmp_repair_limits_xlsx (
  protheus text,
  cooperat text,
  min_qty numeric,
  max_qty numeric,
  reorder_qty numeric,
  previous_balance numeric,
  raw_data jsonb
) on commit drop;
\\copy tmp_repair_limits_xlsx FROM '{psql_path(workbook_file)}' WITH (FORMAT csv, DELIMITER E'\\t', NULL '\\N')
insert into inventory_item_limits (
  item_id, item_legacy_key, source, min_qty, max_qty, reorder_qty,
  previous_balance, applied, raw_data
)
select distinct on (i.id, x.protheus, x.cooperat)
  i.id,
  i.legacy_key,
  'repair:downloads/estoque_minimo.xlsx',
  x.min_qty,
  x.max_qty,
  x.reorder_qty,
  x.previous_balance,
  false,
  x.raw_data
from tmp_repair_limits_xlsx x
join inventory_items i
  on i.protheus_code = x.protheus
  or i.cooperat_code = x.cooperat
where not exists (
  select 1 from inventory_item_limits existing
  where existing.item_id = i.id
    and existing.source = 'repair:inventory_items.raw_data'
);
COMMIT;
"""
    script = work_dir / "repair_limits.sql"
    script.write_text(sql, encoding="utf-8")
    psql.run_file(script)
    return {"workbook_rows": len(workbook_rows)}


def counts(psql: Psql) -> dict[str, int]:
    sql = """
select 'cooperat_purchase_codes', count(*) from cooperat_purchase_codes
union all select 'cooperat_purchase_events', count(*) from cooperat_purchase_events
union all select 'inventory_balance_history', count(*) from inventory_balance_history
union all select 'inventory_item_limits', count(*) from inventory_item_limits
union all select 'inventory_movements', count(*) from inventory_movements
union all select 'inventory_items', count(*) from inventory_items
union all select 'inventory_snapshots', count(*) from inventory_snapshots
order by 1
"""
    result: dict[str, int] = {}
    for row in psql.scalar_rows(sql):
        name, value = row.split("\t", 1)
        result[name] = int(value)
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description="Repara extracoes SQL recuperando dados de arquivos e snapshots antigos.")
    parser.add_argument("--database-url", default=DEFAULT_DATABASE_URL)
    parser.add_argument("--work-dir", default=str(ROOT / "_db_repair_artifacts"))
    parser.add_argument("--cooperat-json", default=str(ROOT / "data" / "historico_cooperat_antigo.json"))
    parser.add_argument("--estoque-minimo", default=str(ROOT / "downloads" / "estoque_minimo.xlsx"))
    args = parser.parse_args()

    work_dir = Path(args.work_dir).resolve() / datetime.now().strftime("repair_%Y%m%d_%H%M%S")
    work_dir.mkdir(parents=True, exist_ok=True)
    psql = Psql(args.database_url)

    before = counts(psql)
    cooperat = repair_cooperat(work_dir, Path(args.cooperat_json), psql)
    history = repair_history(work_dir, psql)
    limits = repair_limits(work_dir, Path(args.estoque_minimo), psql)
    after = counts(psql)
    report = {
        "started_at": work_dir.name,
        "before": before,
        "after": after,
        "cooperat": cooperat,
        "history": history,
        "limits": limits,
        "work_dir": str(work_dir),
    }
    (work_dir / "RELATORIO_REPARO.json").write_text(json_text(report), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
