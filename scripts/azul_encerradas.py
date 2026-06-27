import csv
import logging
import subprocess
import time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ENCERRADAS_SHEET = "REQUISICOES ENCERRADAS"


def _normalize_digits(value: Any, size: int) -> str:
    digits = "".join(ch for ch in str(value or "") if ch.isdigit())
    return digits.zfill(size) if digits else ""


def _key(sa: Any, item: Any) -> str:
    sa_code = _normalize_digits(sa, 6)
    item_code = _normalize_digits(item, 2)
    return f"{sa_code};{item_code}" if sa_code and item_code else ""


def _candidate_mata185_paths(project_root: Path) -> list[Path]:
    return [
        project_root / "downloads" / "mata185.xlsx",
        project_root / "mata185.xlsx",
        Path.home() / "Desktop" / "mata185.xlsx",
        Path.home() / "Downloads" / "mata185.xlsx",
    ]


def _find_mata185(project_root: Path) -> Path:
    candidates = [path for path in _candidate_mata185_paths(project_root) if path.exists()]
    if not candidates:
        raise FileNotFoundError("mata185.xlsx nao encontrada para registrar requisicoes encerradas.")
    return max(candidates, key=lambda path: path.stat().st_mtime)


def _read_keys(path: Path) -> set[str]:
    if not path.exists():
        return set()
    keys: set[str] = set()
    for line in path.read_text(encoding="utf-8-sig", errors="replace").splitlines():
        text = line.strip()
        if not text or "Nr.S" in text:
            continue
        if ";" in text:
            parts = next(csv.reader([text], delimiter=";"))
            if len(parts) >= 2:
                key = _key(parts[0], parts[1])
                if key:
                    keys.add(key)
        else:
            sa = _normalize_digits(text, 6)
            if sa:
                keys.add(f"{sa};")
    return keys


def _write_encerradas_sheet(mata185_path: Path, azul_keys: set[str]) -> dict[str, int]:
    wb = load_workbook(mata185_path)
    try:
        source = wb[wb.sheetnames[0]]
        if ENCERRADAS_SHEET in wb.sheetnames:
            del wb[ENCERRADAS_SHEET]
        target = wb.create_sheet(ENCERRADAS_SHEET)

        max_col = source.max_column
        for col in range(1, max_col + 1):
            target.cell(row=1, column=col).value = source.cell(row=1, column=col).value
        target.cell(row=1, column=max_col + 1).value = "Status Encerramento"
        target.cell(row=1, column=max_col + 2).value = "Chave Azul"

        matched = 0
        rows_seen = 0
        target_row = 2
        for row in range(2, source.max_row + 1):
            rows_seen += 1
            chave = _key(source.cell(row=row, column=1).value, source.cell(row=row, column=2).value)
            if not chave:
                continue
            matches = chave in azul_keys or f"{chave.split(';', 1)[0]};" in azul_keys
            if not matches:
                continue
            for col in range(1, max_col + 1):
                target.cell(row=target_row, column=col).value = source.cell(row=row, column=col).value
            target.cell(row=target_row, column=max_col + 1).value = "Requisicao encerrada"
            target.cell(row=target_row, column=max_col + 2).value = chave
            matched += 1
            target_row += 1

        wb.save(mata185_path)
        return {
            "azuis": len(azul_keys),
            "linhas_mata185": rows_seen,
            "encerradas": matched,
        }
    finally:
        wb.close()


def registrar_requisicoes_encerradas(
    logger: logging.Logger,
    base: Path,
    project_root: Path,
    timeout_seconds: int = 900,
) -> dict[str, int]:
    script = base / "procurar_azul.ps1"
    output_dir = project_root / "downloads"
    output_dir.mkdir(parents=True, exist_ok=True)
    lista_path = output_dir / "lista-azuis.txt"

    if not script.exists():
        raise FileNotFoundError(f"Verificador de azul nao encontrado: {script}")

    logger.info("AZUL: iniciando verificador de requisicoes encerradas.")
    started = time.time()
    subprocess.run(
        [
            "powershell.exe",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(script),
            "-OutputDir",
            str(output_dir),
        ],
        check=True,
        timeout=timeout_seconds,
    )
    logger.info("AZUL: verificador concluido em %.1fs.", time.time() - started)

    keys = _read_keys(lista_path)
    mata185_path = _find_mata185(project_root)
    result = _write_encerradas_sheet(mata185_path, keys)
    logger.info(
        "AZUL: aba %s atualizada em %s | azuis=%s | encerradas=%s",
        ENCERRADAS_SHEET,
        mata185_path,
        result["azuis"],
        result["encerradas"],
    )
    return result
