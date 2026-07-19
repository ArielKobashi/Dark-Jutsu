import argparse
import base64
import hashlib
import json
import logging
import math
import os
import re
import shutil
import sys
import time
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen
from zipfile import BadZipFile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook

try:
    from atualizacao.automus_crypto import decrypt_config, read_json
except ModuleNotFoundError:
    from automus_crypto import decrypt_config, read_json

FILE_MAP = {
    "mata105": "incluir.xlsx",
    "mata225": "Saldo Atual.xlsx",
    "mata226": "Saldo por Endereco.xlsx",
    "estoque_minimo": "estoque_minimo.xlsx",
    "dados_mortos": "dados.mortos.xlsx",
    "mata110": "mata110.xlsx",
    "mata111": "mata111.xlsx",
    "mata112": "mata112.xlsx",
    "mata185": "mata185.xlsx",
}


def _candidate_dashboard_roots(project_root: Path) -> list[Path]:
    roots: list[Path] = []
    env_root = os.environ.get("AUTOMUS_DASHBOARD_ROOT")
    if env_root:
        roots.append(Path(env_root))
    roots.append(Path.home() / "Desktop" / "KKKKKKKK")
    for parent in project_root.resolve().parents:
        if (parent / "dashboard.html").exists():
            roots.append(parent)
            break

    unique: list[Path] = []
    seen: set[str] = set()
    for root in roots:
        key = str(root.resolve()) if root.exists() else str(root)
        if key not in seen:
            seen.add(key)
            unique.append(root)
    return unique


def _mirror_mata112_to_dashboards(mata112_path: Path | None, project_root: Path, log: logging.Logger) -> None:
    if not mata112_path or not mata112_path.exists():
        return

    copied = 0
    for root in _candidate_dashboard_roots(project_root):
        if not (root / "dashboard.html").exists():
            continue
        destino = root / "downloads" / "mata112.xlsx"
        try:
            if destino.exists() and mata112_path.resolve() == destino.resolve():
                continue
            destino.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(mata112_path, destino)
            copied += 1
            log.info("MATA112_DASHBOARD_SYNC_OK | origem=%s | destino=%s", mata112_path, destino)
        except OSError as exc:
            log.warning("MATA112_DASHBOARD_SYNC_FALHOU | origem=%s | destino=%s | erro=%s", mata112_path, destino, exc)

    if not copied:
        log.info("MATA112_DASHBOARD_SYNC_IGNORADO | nenhum dashboard local encontrado")


def _json_hash(value: Any) -> str:
    raw = _json_payload(value, sort_keys=True)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


def _sanitize_json_value(value: Any) -> Any:
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    if isinstance(value, dict):
        return {str(k): _sanitize_json_value(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_sanitize_json_value(v) for v in value]
    return value


def _json_payload(value: Any, sort_keys: bool = False, indent: int | None = None) -> str:
    return json.dumps(
        _sanitize_json_value(value),
        ensure_ascii=False,
        sort_keys=sort_keys,
        separators=None if indent is not None else (",", ":"),
        indent=indent,
        allow_nan=False,
    )


def _http_json(
    url: str,
    method: str = "GET",
    payload: dict[str, Any] | None = None,
    timeout: float = 30.0,
    headers_extra: dict[str, str] | None = None,
) -> dict[str, Any] | list[Any] | None:
    data = None
    headers = {"Accept": "application/json"}
    if headers_extra:
        headers.update(headers_extra)
    if payload is not None:
        data = _json_payload(payload).encode("utf-8")
        headers["Content-Type"] = "application/json"
    req = Request(url=url, data=data, method=method, headers=headers)
    try:
        with urlopen(req, timeout=timeout) as resp:
            raw = resp.read()
            if not raw:
                return None
            return json.loads(raw.decode("utf-8"))
    except HTTPError as exc:
        body = ""
        try:
            body = exc.read().decode("utf-8", errors="replace")
        except Exception:
            pass
        raise RuntimeError(f"HTTP {exc.code} em {url}: {body}") from exc
    except URLError as exc:
        raise RuntimeError(f"Falha de rede em {url}: {exc}") from exc


def _api_base_candidates() -> list[str]:
    candidates = [
        os.environ.get("DARK_JUTSU_API_BASE_URL", ""),
        "http://127.0.0.1:8765",
        "http://192.168.5.44:8765",
        "http://192.168.5.41:8765",
        "http://192.168.5.38:8765",
    ]
    unique: list[str] = []
    seen: set[str] = set()
    for candidate in candidates:
        base_url = str(candidate or "").strip().rstrip("/")
        if not base_url or base_url in seen:
            continue
        seen.add(base_url)
        unique.append(base_url)
    return unique


def _post_inventory_to_sql_api(payload: dict[str, Any], id_token: str, log: logging.Logger) -> dict[str, Any] | None:
    service_token = os.environ.get("DARK_JUTSU_API_TOKEN", "").strip()
    bearer = service_token or id_token
    if not bearer:
        log.warning("AUTOMUS_SQL_IGNORADO | token ausente")
        return None

    last_error: Exception | None = None
    for base_url in _api_base_candidates():
        try:
            result = _http_json(
                f"{base_url}/api/inventory/automus-update",
                method="POST",
                payload=payload,
                timeout=120.0,
                headers_extra={"Authorization": f"Bearer {bearer}"},
            )
            if not isinstance(result, dict) or not result.get("ok"):
                raise RuntimeError(f"Resposta invalida da API SQL: {result}")
            log.info(
                "AUTOMUS_SQL_UPDATE_OK | base=%s | snapshot=%s | itens=%s | enderecos=%s",
                base_url,
                result.get("snapshot_id"),
                result.get("items_loaded"),
                result.get("addresses_loaded"),
            )
            return result
        except Exception as exc:
            last_error = exc
            log.warning("AUTOMUS_SQL_TENTATIVA_FALHOU | base=%s | motivo=%s", base_url, exc)

    raise RuntimeError(f"Nenhuma API SQL aceitou a carga Automus. Ultimo erro: {last_error}")


def _automus_sql_only_enabled() -> bool:
    return True


def _load_automus_config_sql_only(config_path: Path) -> tuple[dict[str, Any], Path | None]:
    if config_path.exists():
        cfg = read_json(config_path)
        if cfg.get("format") == "automus-config-v1":
            raise RuntimeError("AUTOMUS_SQL_ONLY nao abre automus_config criptografado legado. Use DARK_JUTSU_API_TOKEN ou config plano.")
        return cfg, config_path
    return {}, None


def _decode_jwt_payload(token: str) -> dict[str, Any]:
    parts = str(token or "").split(".")
    if len(parts) < 2:
        return {}
    raw = parts[1] + "=" * (-len(parts[1]) % 4)
    try:
        decoded = base64.urlsafe_b64decode(raw.encode("ascii"))
        payload = json.loads(decoded.decode("utf-8"))
        return payload if isinstance(payload, dict) else {}
    except Exception:
        return {}


def _id_token_expired(token: str, skew_seconds: int = 120) -> bool:
    payload = _decode_jwt_payload(token)
    exp = payload.get("exp")
    try:
        return int(exp) <= int(time.time()) + skew_seconds
    except Exception:
        return False


def _load_automus_config(config_path: Path, api_key: str, db_url: str) -> tuple[dict[str, Any], Path | None]:
    if config_path.exists():
        cfg = read_json(config_path)
        if cfg.get("format") == "automus-config-v1":
            return decrypt_config(cfg, api_key, db_url), config_path
        return cfg, config_path

    encrypted_path = config_path.with_name("automus_config.enc.json")
    if encrypted_path.exists():
        return decrypt_config(read_json(encrypted_path), api_key, db_url), encrypted_path

    return {}, None


def _sheet_name_key(value: Any) -> str:
    txt = _safe_text(value).lower()
    normalized = unicodedata.normalize("NFKD", txt)
    return normalized.encode("ascii", "ignore").decode("ascii").strip()


def _read_sheet(path: Path, sheet_name: str | None = None) -> list[list[Any]]:
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        try:
            ws_name = wb.sheetnames[0]
            if sheet_name:
                if sheet_name in wb.sheetnames:
                    ws_name = sheet_name
                else:
                    alvo = _sheet_name_key(sheet_name)
                    ws_name_match = next(
                        (
                            name for name in wb.sheetnames
                            if _sheet_name_key(name) == alvo
                            or alvo in _sheet_name_key(name)
                            or _sheet_name_key(name) in alvo
                        ),
                        None,
                    )
                    if not ws_name_match:
                        raise RuntimeError(
                            f"Aba {sheet_name!r} nao encontrada em {path}. Abas disponiveis: {', '.join(wb.sheetnames)}"
                        )
                    ws_name = ws_name_match
            ws = wb[ws_name]
            out: list[list[Any]] = []
            for row in ws.iter_rows(values_only=True):
                out.append(list(row))
            return out
        finally:
            wb.close()
    except BadZipFile:
        return _read_spreadsheetml_sheet(path, sheet_name)


def _read_spreadsheetml_sheet(path: Path, sheet_name: str | None = None) -> list[list[Any]]:
    ns = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}
    tree = ET.parse(path)
    root = tree.getroot()
    worksheets = root.findall(".//ss:Worksheet", ns)
    if not worksheets:
        return []
    worksheet = worksheets[0]
    if sheet_name:
        alvo = _sheet_name_key(sheet_name)
        worksheet_match = next(
            (
                ws for ws in worksheets
                if alvo in _sheet_name_key(ws.attrib.get("{urn:schemas-microsoft-com:office:spreadsheet}Name", ""))
                or _sheet_name_key(ws.attrib.get("{urn:schemas-microsoft-com:office:spreadsheet}Name", "")) in alvo
            ),
            None,
        )
        if worksheet_match is None:
            nomes = [
                ws.attrib.get("{urn:schemas-microsoft-com:office:spreadsheet}Name", "")
                for ws in worksheets
            ]
            raise RuntimeError(
                f"Aba {sheet_name!r} nao encontrada em {path}. Abas disponiveis: {', '.join(nomes)}"
            )
        worksheet = worksheet_match
    table = worksheet.find(".//ss:Table", ns)
    if table is None:
        return []
    rows: list[list[Any]] = []
    for row_el in table.findall("ss:Row", ns):
        row: list[Any] = []
        for cell in row_el.findall("ss:Cell", ns):
            index_raw = cell.attrib.get("{urn:schemas-microsoft-com:office:spreadsheet}Index")
            if index_raw:
                try:
                    while len(row) < int(index_raw) - 1:
                        row.append(None)
                except ValueError:
                    pass
            data = cell.find("ss:Data", ns)
            row.append(data.text if data is not None else None)
        rows.append(row)
    return rows


def _header_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", _sheet_name_key(value))


def _find_header_index(rows: list[list[Any]] | None, required: tuple[str, ...]) -> int:
    required_keys = {_header_key(name) for name in required}
    for idx, row in enumerate(rows or []):
        keys = {_header_key(value) for value in row if _safe_text(value)}
        if required_keys.issubset(keys):
            return idx
    return 1 if rows and len(rows) > 1 else 0


def _column_map(
    rows: list[list[Any]] | None,
    required: tuple[str, ...],
    aliases: dict[str, tuple[str, ...]],
) -> tuple[int, dict[str, int]]:
    header_index = _find_header_index(rows, required)
    header = (rows or [[]])[header_index] if rows else []
    by_key = {_header_key(value): index for index, value in enumerate(header) if _safe_text(value)}
    out: dict[str, int] = {}
    for name, names in aliases.items():
        for alias in names:
            key = _header_key(alias)
            if key in by_key:
                out[name] = by_key[key]
                break
    return header_index, out


def _cell(row: list[Any], columns: dict[str, int], name: str, fallback: int | None = None) -> Any:
    index = columns.get(name, fallback)
    if index is None or len(row) <= index:
        return ""
    return row[index]


def _is_endereco_valido(value: Any) -> bool:
    txt = ("" if value is None else str(value)).strip()
    if not txt:
        return False
    lower = txt.lower()
    if not re.search(r"[A-Za-zÀ-ÿ]", txt):
        return False
    invalid = {"sem endereco", "sem endereço", "nd", "n/d", "na", "n/a", "-", "--", "/", "/ /"}
    if lower in invalid:
        return False
    return True


def _num(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    txt = str(value).strip()
    if not txt:
        return 0.0
    txt = txt.replace(".", "").replace(",", ".")
    try:
        return float(txt)
    except ValueError:
        return 0.0


def _optional_num(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    txt = str(value).strip()
    if not txt:
        return None
    txt = txt.replace(".", "").replace(",", ".")
    try:
        return float(txt)
    except ValueError:
        return None


def _optional_limit_num(value: Any) -> float | None:
    num = _optional_num(value)
    if num is None or num == 0:
        return None
    return num


def _normalizar_limite_cooperat(value: Any, codigo: Any = "") -> float | None:
    num = _optional_limit_num(value)
    if num is None:
        return None
    codigo_norm = re.sub(r"\D+", "", str(codigo or "")).lstrip("0") or str(codigo or "").strip().lower()
    return num / 1000 if codigo_norm == "104739" and num >= 1000 and num % 1000 == 0 else num


def _reposicao_media(minimo: Any, maximo: Any) -> float | None:
    min_num = _optional_limit_num(minimo)
    max_num = _optional_limit_num(maximo)
    if min_num is None or max_num is None:
        return None
    if max_num <= min_num:
        return None
    return max_num - min_num


def _ceil_positive(value: Any) -> int | None:
    num = _optional_num(value)
    if num is None or num <= 0:
        return None
    return max(1, int(math.ceil(num)))


def _historico_consumo_stats(item: dict[str, Any], historico: dict[str, Any]) -> dict[str, float]:
    hist = historico.get(_ajuste_key(_item_key(item)))
    saidas: list[float] = []
    if isinstance(hist, list):
        for ent in hist[-12:]:
            if not isinstance(ent, dict):
                continue
            delta = _optional_num(ent.get("delta"))
            if delta is not None and delta < 0:
                saidas.append(abs(delta))
    if not saidas:
        return {"media": 0.0, "pico": 0.0, "desvio": 0.0, "eventos": 0.0}
    media = sum(saidas) / len(saidas)
    variancia = sum((v - media) ** 2 for v in saidas) / len(saidas)
    return {
        "media": media,
        "pico": max(saidas),
        "desvio": math.sqrt(variancia),
        "eventos": float(len(saidas)),
    }


def _sugerir_limites_estoque(
    item: dict[str, Any],
    historico: dict[str, Any],
    pedidos_compra: dict[str, Any],
) -> dict[str, Any] | None:
    consumo = _historico_consumo_stats(item, historico)
    saldo_atual = _optional_num(item.get("saldo")) or 0.0
    limites_cooperat = item.get("limitesCooperat") if isinstance(item.get("limitesCooperat"), dict) else {}
    saldo_anterior = _optional_num(limites_cooperat.get("saldoAnterior") if isinstance(limites_cooperat, dict) else None)
    consumo_entre_planilhas = max(0.0, (saldo_anterior or 0.0) - saldo_atual) if saldo_anterior is not None else 0.0

    codigo = _ascii_lower(item.get("protheus") or item.get("protheusKey") or item.get("cooperat"))
    pedido = pedidos_compra.get(_ajuste_key(codigo)) if codigo else None
    if not isinstance(pedido, dict) and codigo:
        pedido = pedidos_compra.get(codigo)
    media_pedido = _optional_num(pedido.get("mediaPedido")) if isinstance(pedido, dict) else None
    media_pedido = media_pedido or 0.0
    ultimo_pedido = pedido.get("ultimoPedido") if isinstance(pedido, dict) and isinstance(pedido.get("ultimoPedido"), dict) else {}
    ultimo_pedido_base = max(
        _optional_num(ultimo_pedido.get("quantidadeRecebida")) or 0.0,
        _optional_num(ultimo_pedido.get("quantidadePedida")) or 0.0,
    )
    base_pedido_compra = media_pedido if media_pedido > 0 else ultimo_pedido_base
    minimo_por_compra = base_pedido_compra * 0.35 if base_pedido_compra > 0 else 0.0

    candidatos_consumo = [
        consumo["media"],
        consumo["pico"] * 0.65,
        consumo_entre_planilhas,
    ]
    candidatos_consumo = [v for v in candidatos_consumo if math.isfinite(v) and v > 0]
    candidatos = [
        *candidatos_consumo,
        minimo_por_compra,
    ]
    candidatos = [v for v in candidatos if math.isfinite(v) and v > 0]
    if not candidatos:
        return None

    demanda_base = max(candidatos)
    estoque_seguranca = max(consumo["desvio"], demanda_base * 0.35, consumo["pico"] * 0.25, minimo_por_compra)
    minimo_estimado = max(demanda_base + estoque_seguranca, minimo_por_compra) if candidatos_consumo else minimo_por_compra
    minimo = _ceil_positive(minimo_estimado)
    if minimo is None:
        return None
    lote = _ceil_positive(max(base_pedido_compra, demanda_base * 1.5, minimo * 0.8)) or minimo
    maximo = max(minimo + lote, minimo + 1)
    return {
        "minimo": minimo,
        "maximo": maximo,
        "reposicao": maximo - minimo,
        "criterio": {
            "demandaBase": demanda_base,
            "estoqueSeguranca": estoque_seguranca,
            "consumoMedio": consumo["media"],
            "consumoPico": consumo["pico"],
            "eventosConsumo": int(consumo["eventos"]),
            "consumoEntrePlanilhas": consumo_entre_planilhas,
            "mediaPedido": media_pedido,
            "ultimoPedidoQuantidade": ultimo_pedido_base,
            "basePedidoCompra": base_pedido_compra,
            "minimoPorCompra": minimo_por_compra,
            "fontePrincipal": "pedido_compra" if base_pedido_compra > 0 and not candidatos_consumo else "consumo",
        },
    }


def _aplicar_sugestoes_estoque(
    itens: list[dict[str, Any]],
    historico: dict[str, Any],
    pedidos_compra: dict[str, Any],
) -> None:
    for item in itens:
        if not isinstance(item, dict):
            continue
        if item.get("limitesOrigem") in {"manual", "cooperat"}:
            item.pop("sugestaoEstoque", None)
            continue
        sugestao = _sugerir_limites_estoque(item, historico, pedidos_compra)
        if not sugestao:
            item.pop("sugestaoEstoque", None)
            continue
        if item.get("minimo") in (None, ""):
            item["minimo"] = sugestao["minimo"]
            item["minimoOrigem"] = "automatico"
        if item.get("maximo") in (None, ""):
            item["maximo"] = sugestao["maximo"]
            item["maximoOrigem"] = "automatico"
        if item.get("minimoOrigem") == "automatico" or item.get("maximoOrigem") == "automatico":
            item["limitesOrigem"] = item.get("limitesOrigem") or "automatico"
            item["reposicao"] = _reposicao_media(item.get("minimo"), item.get("maximo"))
            item["reposicaoOrigem"] = "automatico"
            item["sugestaoEstoque"] = sugestao["criterio"]


def _ajuste_key(item_key: str) -> str:
    raw = item_key.encode("utf-8")
    return base64.urlsafe_b64encode(raw).decode("ascii").rstrip("=")


def _item_key(item: dict[str, Any]) -> str:
    return _safe_text(item.get("protheusKey") or item.get("protheus"))


def _atualizar_historico_saldo(
    novos: list[dict[str, Any]],
    anteriores: list[dict[str, Any]],
    historico_atual: dict[str, Any],
    agora_ms: int,
) -> dict[str, Any]:
    anterior_por_key: dict[str, dict[str, Any]] = {}
    for item in anteriores:
        if not isinstance(item, dict):
            continue
        key = _item_key(item)
        if key and key not in anterior_por_key:
            anterior_por_key[key] = item

    historico = dict(historico_atual or {})
    data_txt = datetime.fromtimestamp(agora_ms / 1000).strftime("%d/%m/%Y")

    for item in novos:
        if not isinstance(item, dict):
            continue
        key = _item_key(item)
        if not key or key not in anterior_por_key:
            continue
        saldo_anterior = _optional_num(anterior_por_key[key].get("saldo"))
        saldo_atual = _optional_num(item.get("saldo"))
        if saldo_anterior is None or saldo_atual is None:
            continue
        delta = saldo_atual - saldo_anterior
        if not delta:
            continue
        hist_key = _ajuste_key(key)
        lista = historico.get(hist_key) if isinstance(historico.get(hist_key), list) else []
        lista = list(lista)
        lista.append(
            {
                "data": data_txt,
                "timestamp": agora_ms,
                "delta": delta,
                "tipo": "entrada" if delta > 0 else "saida",
                "saldoAnterior": saldo_anterior,
                "saldoAtual": saldo_atual,
            }
        )
        historico[hist_key] = lista[-300:]

    return historico


def _sheet_date(value: Any) -> tuple[str, int]:
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y"), int(value.timestamp() * 1000)
    txt = _safe_text(value)
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(txt[:10], fmt)
            return dt.strftime("%d/%m/%Y"), int(dt.timestamp() * 1000)
        except Exception:
            pass
    return txt, 0


def _entrada_depois(item: dict[str, Any], historico: dict[str, Any], timestamp: int) -> bool:
    hist = historico.get(_ajuste_key(_item_key(item)))
    if not isinstance(hist, list):
        return False
    for ent in hist:
        if not isinstance(ent, dict):
            continue
        if ent.get("tipo") == "entrada" and _num(ent.get("delta")) > 0 and (not timestamp or int(ent.get("timestamp") or 0) >= timestamp):
            return True
    return False


def _montar_pedidos_compra(
    mata110: list[list[Any]] | None,
    mata111: list[list[Any]] | None,
    mata112: list[list[Any]] | None,
    itens: list[dict[str, Any]],
    historico: dict[str, Any],
) -> dict[str, Any]:
    mata111_header, mata111_cols = _column_map(mata111, ("Produto",), {
        "codigo": ("Produto",),
        "data": ("Data Emissao", "Data Emissão"),
        "quantidade_pedida": ("Quantidade",),
        "quantidade_recebida": ("Qtd.Entregue", "Qtd Entregue"),
    })
    pedidos_por_codigo: dict[str, list[dict[str, Any]]] = {}
    for idx, row in enumerate((mata111 or [])[mata111_header + 1:]):
        codigo = _ascii_lower(_cell(row, mata111_cols, "codigo", 6))
        if not codigo:
            continue
        data_txt, ts = _sheet_date(_cell(row, mata111_cols, "data", 2))
        pedidos_por_codigo.setdefault(codigo, []).append({
            "codigo": codigo,
            "dataPedido": data_txt,
            "timestamp": ts,
            "quantidadePedida": _num(_cell(row, mata111_cols, "quantidade_pedida", 9)),
            "quantidadeRecebida": _num(_cell(row, mata111_cols, "quantidade_recebida", 11)),
            "rowIndex": idx,
        })
    for lista in pedidos_por_codigo.values():
        lista.sort(key=lambda p: (p.get("timestamp") or 0, p.get("rowIndex") or 0), reverse=True)

    mata110_header, mata110_cols = _column_map(mata110, ("Produto",), {
        "codigo": ("Produto",),
        "quantidade_solicitada": ("Quantidade",),
        "solicitante": ("Solicitante",),
        "quantidade_em_pedido": ("Quant.em Ped", "Quant em Ped"),
        "aceito_por": ("Nome",),
    })
    solicitacoes_por_codigo: dict[str, list[dict[str, Any]]] = {}
    for idx, row in enumerate((mata110 or [])[mata110_header + 1:]):
        codigo = _ascii_lower(_cell(row, mata110_cols, "codigo", 3))
        if not codigo:
            continue
        solicitacoes_por_codigo.setdefault(codigo, []).append({
            "codigo": codigo,
            "quantidadeSolicitada": _num(_cell(row, mata110_cols, "quantidade_solicitada", 5)),
            "solicitante": _safe_text(_cell(row, mata110_cols, "solicitante", 7)),
            "quantidadeEmPedido": _num(_cell(row, mata110_cols, "quantidade_em_pedido", 8)),
            "aceitoPor": _safe_text(_cell(row, mata110_cols, "aceito_por", 17)),
            "rowIndex": idx,
        })

    mata112_header, mata112_cols = _column_map(mata112, ("Produto",), {
        "codigo": ("Produto",),
        "quantidade_entrada": ("Quantidade", "Quantidade Original", "Quantidade_x000D_\nOriginal"),
        "saldo_a_enderecar": ("Saldo a Enderecar", "Saldo a Endereçar", "Saldo a Distribuir", "Saldo a_x000D_\nDistribuir"),
        "data_entrada": ("Data", "Dt Digitacao", "Dt Digitação", "Data Digitacao", "Data Digitação"),
    })
    entradas_por_codigo: dict[str, list[dict[str, Any]]] = {}
    for idx, row in enumerate((mata112 or [])[mata112_header + 1:]):
        codigo = _ascii_lower(_cell(row, mata112_cols, "codigo", 0))
        if not codigo:
            continue
        data_txt, ts = _sheet_date(_cell(row, mata112_cols, "data_entrada", 12))
        entradas_por_codigo.setdefault(codigo, []).append({
            "codigo": codigo,
            "quantidadeEntrada": _num(_cell(row, mata112_cols, "quantidade_entrada", 3)),
            "saldoAEnderecar": _num(_cell(row, mata112_cols, "saldo_a_enderecar", 4)),
            "dataEntrada": data_txt,
            "timestamp": ts,
            "rowIndex": idx,
        })
    for lista in entradas_por_codigo.values():
        lista.sort(key=lambda e: (e.get("timestamp") or 0, e.get("rowIndex") or 0), reverse=True)

    estado: dict[str, Any] = {}
    for item in itens:
        codigo = _ascii_lower(item.get("protheus") or item.get("protheusKey") or item.get("cooperat"))
        if not codigo:
            continue
        pedidos = pedidos_por_codigo.get(codigo, [])
        solicitacoes = solicitacoes_por_codigo.get(codigo, [])
        entradas = entradas_por_codigo.get(codigo, [])
        entrada_ativa = None

        def localizar_entrada(pedido: dict[str, Any]) -> dict[str, Any] | None:
            pedido_ts = int(pedido.get("timestamp") or 0)
            return next((e for e in entradas if not pedido_ts or not int(e.get("timestamp") or 0) or int(e.get("timestamp") or 0) >= pedido_ts), None)

        entrada_confirmada = next((e for e in (localizar_entrada(p) for p in pedidos) if e), None) or (entradas[0] if entradas else None)
        pedido = None
        for candidato in pedidos:
            entrada = localizar_entrada(candidato)
            if entrada and _num(entrada.get("saldoAEnderecar")) <= 0:
                continue
            if not entrada and _entrada_depois(item, historico, int(candidato.get("timestamp") or 0)):
                continue
            pedido = candidato
            entrada_ativa = entrada
            break
        solicitacao = next(reversed(solicitacoes), None) if solicitacoes else None
        recebidos = [p for p in pedidos if _num(p.get("quantidadeRecebida")) > 0]
        ultimo_pedido = pedidos[0] if pedidos else None
        media = sum(_num(p.get("quantidadeRecebida")) or _num(p.get("quantidadePedida")) for p in recebidos) / len(recebidos) if recebidos else None
        titulo, detalhe, status, recebido_sem_entrada = "Sem solicitação de pedido", "", "sem", 0
        if solicitacao:
            if not solicitacao.get("aceitoPor") and _num(solicitacao.get("quantidadeEmPedido")) <= 0:
                status, titulo, detalhe = "aguardando_liberacao", "Aguardando liberação...", f"Solicitado por: {solicitacao.get('solicitante') or '-'}"
            elif not pedido and solicitacao.get("aceitoPor"):
                status, titulo, detalhe = "aguardando_pedido", "Aguardando pedido de compra...", f"Aceito por: {solicitacao.get('aceitoPor') or '-'}"
            elif pedido and _num(pedido.get("quantidadeRecebida")) <= 0:
                status, titulo, detalhe = "aguardando_recebimento", "Aguardando recebimento...", f"Pedido feito em: {pedido.get('dataPedido') or '-'}"
            elif pedido and entrada_ativa and _num(entrada_ativa.get("saldoAEnderecar")) > 0:
                status = "aguardando_enderecamento"
                titulo = "Aguardando endereçamento..."
                detalhe = f"Entrou dia: {entrada_ativa.get('dataEntrada') or '-'} - {_format_num(entrada_ativa.get('quantidadeEntrada'))} unidades"
            elif pedido:
                status, titulo, detalhe = "aguardando_contagem", "Aguardando contagem de nota...", f"Recebido em: {pedido.get('dataPedido') or '-'}"
                recebido_sem_entrada = _num(pedido.get("quantidadeRecebida"))
        estado[_ajuste_key(codigo)] = {
            "codigo": codigo,
            "status": status,
            "titulo": titulo,
            "detalhe": detalhe,
            "solicitacao": solicitacao,
            "pedido": pedido,
            "ultimoPedido": ultimo_pedido,
            "entradaEndereco": entrada_ativa,
            "entradaConfirmada": entrada_confirmada,
            "mediaPedido": media,
            "recebidoSemEntrada": recebido_sem_entrada,
        }
    return estado


def _validar_entradas_mata112_refletidas(
    mata112: list[list[Any]] | None,
    itens: list[dict[str, Any]],
    pedidos_compra: dict[str, Any],
    log: logging.Logger,
) -> None:
    if not mata112:
        return

    mata112_header, mata112_cols = _column_map(mata112, ("Produto",), {
        "codigo": ("Produto",),
        "quantidade_entrada": ("Quantidade", "Quantidade Original", "Quantidade_x000D_\nOriginal"),
        "data_entrada": ("Data", "Dt Digitacao", "Dt Digitação", "Data Digitacao", "Data Digitação"),
    })
    entradas_por_data: dict[int, set[str]] = {}
    for row in (mata112 or [])[mata112_header + 1:]:
        codigo = _ascii_lower(_cell(row, mata112_cols, "codigo", 0))
        if not codigo:
            continue
        if _num(_cell(row, mata112_cols, "quantidade_entrada", 3)) <= 0:
            continue
        _data_txt, ts = _sheet_date(_cell(row, mata112_cols, "data_entrada", 12))
        if not ts:
            continue
        entradas_por_data.setdefault(ts, set()).add(codigo)

    if not entradas_por_data:
        return

    data_mais_recente = max(entradas_por_data)
    codigos_entrada = entradas_por_data[data_mais_recente]
    codigos_itens = {
        _ascii_lower(item.get("protheus") or item.get("protheusKey") or item.get("cooperat"))
        for item in itens
        if isinstance(item, dict)
    }
    codigos_esperados = {codigo for codigo in codigos_entrada if codigo in codigos_itens}
    if not codigos_esperados:
        log.warning(
            "VALIDACAO_MATA112_SEM_ITENS_CADASTRADOS | data=%s | codigos=%s",
            datetime.fromtimestamp(data_mais_recente / 1000).strftime("%d/%m/%Y"),
            ",".join(sorted(codigos_entrada)),
        )
        return

    faltando: list[str] = []
    for codigo in sorted(codigos_esperados):
        pedido = pedidos_compra.get(_ajuste_key(codigo)) or pedidos_compra.get(codigo)
        if not isinstance(pedido, dict):
            faltando.append(codigo)
            continue
        entrada = pedido.get("entradaConfirmada") or pedido.get("entradaEndereco")
        if not isinstance(entrada, dict):
            faltando.append(codigo)

    if faltando:
        data_br = datetime.fromtimestamp(data_mais_recente / 1000).strftime("%d/%m/%Y")
        raise RuntimeError(
            "NAO CONFORME: MATA112 possui entradas recentes que nao foram refletidas em pedidosCompra "
            f"(data={data_br}; faltando={', '.join(faltando[:80])}"
            f"{'...' if len(faltando) > 80 else ''}). "
            "Atualizacao interrompida para evitar Kanban de entrada defasado."
        )

    log.info(
        "VALIDACAO_MATA112_ENTRADAS_OK | data=%s | codigos=%s",
        datetime.fromtimestamp(data_mais_recente / 1000).strftime("%d/%m/%Y"),
        len(codigos_esperados),
    )


def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _format_num(value: Any) -> str:
    num = _num(value)
    if float(num).is_integer():
        return str(int(num))
    return f"{num:.2f}".replace(".", ",")


def _ascii_lower(value: Any) -> str:
    txt = _safe_text(value).lower()
    if not txt:
        return ""
    normalized = unicodedata.normalize("NFKD", txt)
    return normalized.encode("ascii", "ignore").decode("ascii")


def _code_key(value: Any) -> str:
    txt = _safe_text(value).replace(" ", "").lower()
    if not txt:
        return ""
    normalized_num = txt.replace(",", ".")
    if re.fullmatch(r"\d+(?:\.0+)?", normalized_num):
        return str(int(float(normalized_num)))
    return txt


def _req_item_key(requisicao: Any, sequencia: Any) -> str:
    req = re.sub(r"\D+", "", _safe_text(requisicao))
    seq = re.sub(r"\D+", "", _safe_text(sequencia))
    if not req or not seq:
        return ""
    return f"{req.zfill(6)};{seq.zfill(2)}"


def _dead_address_from_row(row: list[Any]) -> str:
    endereco = _safe_text(row[3] if len(row) > 3 else "")
    return endereco if _is_endereco_valido(endereco) else ""


def _dead_addresses_index(rows: list[list[Any]] | None) -> dict[str, str]:
    idx: dict[str, str] = {}
    for row in rows or []:
        if not row:
            continue
        cooperat = _code_key(row[0] if len(row) > 0 else "")
        if not cooperat or "cooperat" in cooperat or "codigo" in cooperat:
            continue
        endereco = _dead_address_from_row(row)
        if endereco and cooperat not in idx:
            idx[cooperat] = endereco
    return idx


def _montar_movimentacoes_mata185(
    mata185: list[list[Any]] | None,
    mata185_encerradas: list[list[Any]] | None,
    itens: list[dict[str, Any]],
    historico_atual: dict[str, Any] | None,
    agora_ms: int,
) -> dict[str, Any]:
    item_por_codigo: dict[str, dict[str, Any]] = {}
    for item in itens:
        if not isinstance(item, dict):
            continue
        for campo in ("protheus", "protheusKey"):
            codigo = _code_key(item.get(campo))
            if codigo and codigo not in item_por_codigo:
                item_por_codigo[codigo] = item

    agregados: dict[str, dict[str, Any]] = {}
    parciais: list[dict[str, Any]] = []
    encerradas: list[dict[str, Any]] = []
    linhas_hash: list[dict[str, Any]] = []
    encerradas_keys = {
        key for key in (
            _req_item_key(row[0] if len(row) > 0 else "", row[1] if len(row) > 1 else "")
            for row in (mata185_encerradas or [])[1:]
            if row
        )
        if key
    }

    for idx, row in enumerate((mata185 or [])[1:], start=2):
        if not row:
            continue
        requisicao = _safe_text(row[0] if len(row) > 0 else "")
        sequencia = _safe_text(row[1] if len(row) > 1 else "")
        codigo = _code_key(row[2] if len(row) > 2 else "")
        unidade = _safe_text(row[3] if len(row) > 3 else "")
        requisitada = _num(row[4] if len(row) > 4 else 0)
        atendida = _num(row[6] if len(row) > 6 else 0)
        if not codigo or "produto" in codigo or "codigo" in codigo:
            continue
        if requisitada <= 0 and atendida <= 0:
            continue

        item = item_por_codigo.get(codigo)
        if item is not None and unidade:
            item["unidadeMedida"] = unidade

        linhas_hash.append({
            "r": requisicao,
            "s": sequencia,
            "c": codigo,
            "u": unidade,
            "q": requisitada,
            "a": atendida,
        })

        ag = agregados.setdefault(codigo, {
            "codigo": codigo,
            "descricao": _safe_text(item.get("descricao")) if item else "",
            "unidade": unidade,
            "requisitado": 0.0,
            "atendido": 0.0,
            "requisicoes": 0,
            "parciais": 0,
        })
        if unidade and not ag.get("unidade"):
            ag["unidade"] = unidade
        ag["requisitado"] += requisitada
        ag["atendido"] += atendida
        ag["requisicoes"] += 1

        if atendida < requisitada:
            faltante = max(0.0, requisitada - atendida)
            payload = {
                "requisicao": requisicao,
                "sequencia": sequencia,
                "codigo": codigo,
                "descricao": ag.get("descricao") or "",
                "unidade": unidade,
                "requisitado": requisitada,
                "atendido": atendida,
                "faltante": faltante,
                "rowIndex": idx,
            }
            if _req_item_key(requisicao, sequencia) in encerradas_keys:
                payload["status"] = "Requisicao encerrada"
                encerradas.append(payload)
            else:
                payload["status"] = "Entregue parcialmente"
                ag["parciais"] += 1
                parciais.append(payload)

    assinatura = _json_hash(linhas_hash)
    data_txt = datetime.fromtimestamp(agora_ms / 1000).strftime("%d/%m/%Y")
    agregados_lista = sorted(agregados.values(), key=lambda x: float(x.get("atendido") or 0), reverse=True)
    parciais = sorted(parciais, key=lambda x: float(x.get("faltante") or 0), reverse=True)
    encerradas = sorted(encerradas, key=lambda x: float(x.get("faltante") or 0), reverse=True)
    historico = dict(historico_atual or {})
    lotes = historico.get("lotes") if isinstance(historico.get("lotes"), list) else []
    lotes = list(lotes)

    lote = {
        "timestamp": agora_ms,
        "data": data_txt,
        "armazem": "04",
        "hash": assinatura,
        "totalItens": len(agregados_lista),
        "totalRequisicoes": sum(int(x.get("requisicoes") or 0) for x in agregados_lista),
        "totalRequisitado": sum(float(x.get("requisitado") or 0) for x in agregados_lista),
        "totalAtendido": sum(float(x.get("atendido") or 0) for x in agregados_lista),
        "totalParciais": len(parciais),
        "totalEncerradas": len(encerradas),
        "itens": agregados_lista[:500],
        "parciais": parciais[:500],
        "encerradas": encerradas[:500],
    }

    if linhas_hash and (not lotes or lotes[-1].get("hash") != assinatura):
        lotes.append(lote)
    elif lotes:
        lotes[-1] = {**lotes[-1], **lote, "timestamp": lotes[-1].get("timestamp") or agora_ms, "data": lotes[-1].get("data") or data_txt}
    elif linhas_hash:
        lotes.append(lote)

    return {
        "armazem": "04",
        "atualizadoEm": agora_ms,
        "loteAtual": lote,
        "lotes": lotes[-120:],
    }


def _detect_incluir_columns(incluir: list[list[Any]]) -> dict[str, int]:
    header = incluir[1] if len(incluir) > 1 else incluir[0] if incluir else []
    cols = {"protheus": 1, "cooperat": 0, "descricao": 2, "inicio": 2}

    for index, value in enumerate(header):
        name = _ascii_lower(value)
        if not name:
            continue
        is_codigo = name in {"codigo", "cod produto", "codigo produto"}
        is_descricao = name in {"descricao", "descrição"}
        is_referencia = "refer" in name or "antigo" in name or name == "cooperat"

        if is_codigo:
            cols["protheus"] = index
        if is_descricao:
            cols["descricao"] = index
        if is_referencia:
            cols["cooperat"] = index

    if len(header) > 1 and _ascii_lower(header[0]) == "codigo" and _ascii_lower(header[1]) == "descricao":
        cols["protheus"] = 0
        cols["descricao"] = 1
        cols["cooperat"] = 2

    return cols


def _is_descricao_dado_morto(value: Any) -> bool:
    desc = _ascii_lower(value)
    return "desativado" in desc or "bloqueado" in desc


def _morto_key(item: dict[str, Any]) -> str:
    return _safe_text(item.get("protheus") or item.get("cooperat") or item.get("protheusKey"))


def _preparar_item_morto(item: dict[str, Any]) -> dict[str, Any]:
    morto = dict(item)
    key = _morto_key(morto)
    morto["morto"] = True
    if key:
        morto["protheusKey"] = "MORTO|" + key
    return morto


def _separar_dados_mortos_por_descricao(
    itens: list[dict[str, Any]],
    dados_mortos_anteriores: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    ativos: list[dict[str, Any]] = []
    mortos: list[dict[str, Any]] = []
    codigos_ativos: set[str] = set()
    codigos_mortos: set[str] = set()

    for item in itens:
        key = _morto_key(item)
        if _is_descricao_dado_morto(item.get("descricao")):
            morto = _preparar_item_morto(item)
            if key and key in codigos_mortos:
                continue
            if key:
                codigos_mortos.add(key)
            mortos.append(morto)
            continue
        ativo = dict(item)
        ativo["morto"] = False
        ativo.pop("protheusKey", None)
        if key:
            codigos_ativos.add(key)
        ativos.append(ativo)

    for item in dados_mortos_anteriores:
        if not isinstance(item, dict):
            continue
        key = _morto_key(item)
        if key and (key in codigos_ativos or key in codigos_mortos):
            continue
        morto = dict(item)
        morto["morto"] = True
        mortos.append(morto)
        if key:
            codigos_mortos.add(key)

    return ativos, mortos


def _build_items(
    incluir: list[list[Any]],
    cooperat: list[list[Any]],
    saldo_atual: list[list[Any]],
    saldo_endereco: list[list[Any]],
    estoque_minimo: list[list[Any]] | None,
    dados_mortos_planilha: list[list[Any]] | None,
    dados_anteriores: list[dict[str, Any]],
    ajustes_itens: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    novos: list[dict[str, Any]] = []
    dados_mortos_endereco_idx = _dead_addresses_index(dados_mortos_planilha)

    saldo_atual_idx: dict[str, float] = {}
    for row in saldo_atual[1:]:
        if len(row) < 5:
            continue
        cod = _code_key(row[1])
        if cod:
            saldo_atual_idx[cod] = _num(row[5] if len(row) > 5 else row[4])

    enderecos_idx: dict[str, list[dict[str, Any]]] = {}
    for row in saldo_endereco[1:]:
        if len(row) < 9:
            continue
        cod = _code_key(row[1])
        if not cod:
            continue
        ent = {
            "endereco": _safe_text(row[4]) or "ND",
            "armazem": _safe_text(row[2]) or "ND",
            "saldo": _num(row[8]),
        }
        enderecos_idx.setdefault(cod, []).append(ent)

    estoque_minimo_idx: dict[str, dict[str, float]] = {}
    for row in (estoque_minimo or [])[1:]:
        if len(row) < 2:
            continue
        cooperat_cod = _safe_text(row[1] if len(row) > 1 else "")
        if not cooperat_cod:
            continue
        cod_lower = _ascii_lower(cooperat_cod)
        if "codigo" in cod_lower or "código" in cod_lower or "item" in cod_lower:
            continue
        minimo = _normalizar_limite_cooperat(row[6] if len(row) > 6 else None, cooperat_cod)
        maximo = _normalizar_limite_cooperat(row[7] if len(row) > 7 else None, cooperat_cod)
        reposicao_planilha = _normalizar_limite_cooperat(row[8] if len(row) > 8 else None, cooperat_cod)
        saldo_anterior = _optional_num(row[5] if len(row) > 5 else None)
        reposicao = reposicao_planilha if reposicao_planilha is not None else _reposicao_media(minimo, maximo)
        if minimo is None and maximo is None and reposicao is None and saldo_anterior is None:
            continue
        estoque_minimo_idx[cooperat_cod] = {
            "temLinha": True,
            "minimo": minimo,
            "maximo": maximo,
            "reposicao": reposicao,
            "saldoAnterior": saldo_anterior,
        }

    prev_by_protheus: dict[str, dict[str, Any]] = {}
    prev_by_cooperat: dict[str, dict[str, Any]] = {}
    prev_by_descricao: dict[str, dict[str, Any]] = {}
    for d in dados_anteriores:
        if not isinstance(d, dict):
            continue
        p = _code_key(d.get("protheus"))
        c = _code_key(d.get("cooperat"))
        desc = _ascii_lower(d.get("descricao"))
        if p and p not in prev_by_protheus:
            prev_by_protheus[p] = d
        if c and c not in prev_by_cooperat:
            prev_by_cooperat[c] = d
        if desc and desc not in prev_by_descricao:
            prev_by_descricao[desc] = d

    incluir_cols = _detect_incluir_columns(incluir)

    for row in incluir[incluir_cols["inicio"]:]:
        if len(row) <= incluir_cols["protheus"]:
            continue

        protheus = _safe_text(row[incluir_cols["protheus"]])
        cooperat_cod = _safe_text(row[incluir_cols["cooperat"]] if len(row) > incluir_cols["cooperat"] else "")
        descricao = _safe_text(row[incluir_cols["descricao"]] if len(row) > incluir_cols["descricao"] else "")
        if not protheus:
            continue
        protheus_key = _code_key(protheus)
        cooperat_key = _code_key(cooperat_cod)

        item = {
            "protheus": protheus,
            "cooperat": cooperat_cod,
            "descricao": descricao,
            "enderecoPrincipal": "Sem Endereco",
            "armazemPrincipal": "ND",
            "saldo": 0.0,
            "enderecos": [],
            "comentarios": [],
            "morto": False,
        }

        if protheus_key in saldo_atual_idx:
            item["saldo"] = saldo_atual_idx[protheus_key]

        enderecos = enderecos_idx.get(protheus_key, [])
        tem_endereco_com_saldo = False
        if enderecos:
            item["enderecos"] = enderecos
            validos = [e for e in enderecos if _is_endereco_valido(e.get("endereco")) and _num(e.get("saldo")) > 0]
            if validos:
                tem_endereco_com_saldo = True
                item["enderecoPrincipal"] = _safe_text(validos[-1].get("endereco"))
                item["armazemPrincipal"] = _safe_text(validos[-1].get("armazem")) or "ND"
            item["saldo"] = sum(_num(e.get("saldo")) for e in enderecos)

        if not tem_endereco_com_saldo:
            endereco_morto = dados_mortos_endereco_idx.get(cooperat_key)
            if endereco_morto:
                item["enderecoPrincipal"] = endereco_morto
                item["armazemPrincipal"] = "ND"

        if not tem_endereco_com_saldo and not _is_endereco_valido(item["enderecoPrincipal"]):
            anterior = prev_by_protheus.get(protheus_key) or prev_by_cooperat.get(cooperat_key) or prev_by_descricao.get(_ascii_lower(descricao))
            if anterior and _is_endereco_valido(anterior.get("enderecoPrincipal")):
                item["enderecoPrincipal"] = _safe_text(anterior.get("enderecoPrincipal"))

        anterior = prev_by_protheus.get(protheus_key) or prev_by_cooperat.get(cooperat_key) or prev_by_descricao.get(_ascii_lower(descricao))
        if anterior and isinstance(anterior.get("enderecos"), list):
            chaves_atuais = {
                (_ascii_lower(e.get("endereco")), _ascii_lower(e.get("armazem")))
                for e in item["enderecos"]
                if isinstance(e, dict)
            }
            for end_antigo in anterior["enderecos"]:
                if not isinstance(end_antigo, dict) or not _is_endereco_valido(end_antigo.get("endereco")):
                    continue
                chave = (_ascii_lower(end_antigo.get("endereco")), _ascii_lower(end_antigo.get("armazem")))
                if chave in chaves_atuais:
                    continue
                item["enderecos"].append({
                    "endereco": _safe_text(end_antigo.get("endereco")),
                    "armazem": _safe_text(end_antigo.get("armazem")) or "ND",
                    "saldo": 0.0,
                    "origem": "Historico",
                })
                chaves_atuais.add(chave)

        estoque_minimo_item = estoque_minimo_idx.get(cooperat_cod)
        if estoque_minimo_item:
            item["limitesCooperat"] = {
                "minimo": estoque_minimo_item.get("minimo"),
                "maximo": estoque_minimo_item.get("maximo"),
                "reposicao": estoque_minimo_item.get("reposicao"),
                "saldoAnterior": estoque_minimo_item.get("saldoAnterior"),
            }
            tem_limite_planilha = (
                estoque_minimo_item.get("minimo") is not None
                or estoque_minimo_item.get("maximo") is not None
                or estoque_minimo_item.get("reposicao") is not None
            )
            if tem_limite_planilha and estoque_minimo_item.get("minimo") is not None:
                item["minimo"] = estoque_minimo_item["minimo"]
                item["minimoOrigem"] = "cooperat"
            if tem_limite_planilha and estoque_minimo_item.get("maximo") is not None:
                item["maximo"] = estoque_minimo_item["maximo"]
                item["maximoOrigem"] = "cooperat"
            if tem_limite_planilha and estoque_minimo_item.get("reposicao") is not None:
                item["reposicao"] = estoque_minimo_item["reposicao"]
                item["reposicaoOrigem"] = "cooperat"
            if tem_limite_planilha:
                item["limitesOrigem"] = "cooperat"

        anterior = prev_by_protheus.get(protheus_key) or prev_by_cooperat.get(cooperat_key) or prev_by_descricao.get(_ascii_lower(descricao))
        if anterior:
            if estoque_minimo is None and not item.get("limitesCooperat") and isinstance(anterior.get("limitesCooperat"), dict):
                item["limitesCooperat"] = anterior.get("limitesCooperat")
            if estoque_minimo is None and item.get("minimo") is None and anterior.get("minimo") is not None:
                item["minimo"] = anterior.get("minimo")
                item["minimoOrigem"] = anterior.get("minimoOrigem") or anterior.get("limitesOrigem") or "anterior"
            if estoque_minimo is None and item.get("maximo") is None and anterior.get("maximo") is not None:
                item["maximo"] = anterior.get("maximo")
                item["maximoOrigem"] = anterior.get("maximoOrigem") or anterior.get("limitesOrigem") or "anterior"
            if estoque_minimo is None and not item.get("limitesOrigem"):
                item["limitesOrigem"] = anterior.get("limitesOrigem") or "anterior"
            if estoque_minimo is None and not item.get("reposicaoOrigem"):
                item["reposicaoOrigem"] = anterior.get("reposicaoOrigem") or item.get("limitesOrigem")

        ajuste = (ajustes_itens or {}).get(_ajuste_key(protheus))
        if isinstance(ajuste, dict):
            if "minimo" in ajuste:
                item["minimo"] = _optional_limit_num(ajuste.get("minimo"))
                item["minimoOrigem"] = "manual"
            if "maximo" in ajuste:
                item["maximo"] = _optional_limit_num(ajuste.get("maximo"))
                item["maximoOrigem"] = "manual"
            item["limitesOrigem"] = "manual"

        reposicao_cooperat = None
        if item.get("limitesOrigem") == "cooperat" and isinstance(item.get("limitesCooperat"), dict):
            reposicao_cooperat = _optional_limit_num(item["limitesCooperat"].get("reposicao"))
        reposicao_final = reposicao_cooperat if reposicao_cooperat is not None else _reposicao_media(item.get("minimo"), item.get("maximo"))
        if reposicao_final is None:
            item.pop("reposicao", None)
        else:
            item["reposicao"] = reposicao_final
            item["reposicaoOrigem"] = item.get("limitesOrigem") or item.get("reposicaoOrigem") or "calculada"

        novos.append(item)

    return novos


def _resolve_file(base_dir: Path, filename: str) -> Path:
    candidates = [
        base_dir / "downloads" / filename,
        base_dir / filename,
        base_dir / "data" / filename,
        Path.home() / "Desktop" / "AMBIENTE ROSA" / filename,
        Path.home() / "Desktop" / filename,
        Path.home() / "Downloads" / filename,
    ]
    for c in candidates:
        if c.exists() and c.is_file():
            return c
    raise FileNotFoundError(f"Arquivo nao encontrado: {filename}")


def _resolve_optional_file(base_dir: Path, filename: str) -> Path | None:
    candidates = [
        base_dir / "downloads" / filename,
        base_dir / filename,
        base_dir / "data" / filename,
        Path.home() / "Desktop" / "AMBIENTE ROSA" / filename,
        Path.home() / "Desktop" / filename,
        Path.home() / "Downloads" / filename,
    ]
    for c in candidates:
        if c.exists() and c.is_file():
            return c
    return None


def _write_local_backup(project_root: Path, backup_payload: dict[str, Any], agora_ms: int, log: logging.Logger) -> Path:
    backup_dir = project_root / "_backups" / "automus"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / f"sql_backup_{agora_ms}.json"
    backup_path.write_text(
        _json_payload(backup_payload, indent=2),
        encoding="utf-8",
    )
    if not backup_path.exists() or backup_path.stat().st_size <= 0:
        raise RuntimeError("NAO CONFORME: backup local nao foi salvo corretamente.")
    log.info(
        "BACKUP_LOCAL_OK | arquivo=%s | bytes=%s",
        backup_path,
        backup_path.stat().st_size,
    )
    return backup_path


def run_automus_update(
    config_path: Path,
    project_root: Path,
    logger: logging.Logger | None = None,
    auth_id_token: str | None = None,
    auth_email: str | None = None,
) -> None:
    log = logger or logging.getLogger("automus_update")
    cfg, resolved_config_path = _load_automus_config_sql_only(config_path)
    service_token = str(auth_id_token or os.environ.get("DARK_JUTSU_API_TOKEN", "")).strip()
    if not resolved_config_path and not service_token:
        raise FileNotFoundError(
            f"Config do Automus nao encontrada em {config_path}. Crie automus_config.json ou automus_config.enc.json."
        )

    email = _safe_text(cfg.get("email"))
    password = _safe_text(cfg.get("password"))
    updated_by = _safe_text(cfg.get("updated_by")) or "atualizado automaticamente via Automus"

    log.info("TEST_AUTOMUS_CONFIG_OK | config=%s | email=%s", resolved_config_path or config_path, auth_email or email)
    if not service_token:
        raise RuntimeError("DARK_JUTSU_API_TOKEN obrigatorio para publicar o Automus no SQL.")

    incluir_path = _resolve_file(project_root, FILE_MAP["mata105"])
    saldo_atual_path = _resolve_file(project_root, FILE_MAP["mata225"])
    saldo_endereco_path = _resolve_file(project_root, FILE_MAP["mata226"])
    estoque_minimo_path = _resolve_optional_file(project_root, FILE_MAP["estoque_minimo"])
    dados_mortos_planilha_path = _resolve_optional_file(project_root, FILE_MAP["dados_mortos"])
    mata110_path = _resolve_optional_file(project_root, FILE_MAP["mata110"])
    mata111_path = _resolve_optional_file(project_root, FILE_MAP["mata111"])
    mata112_path = _resolve_optional_file(project_root, FILE_MAP["mata112"])
    mata185_path = _resolve_optional_file(project_root, FILE_MAP["mata185"])
    log.info(
        "TEST_PLANILHAS_RESOLVIDAS_OK | incluir=%s | saldoAtual=%s | saldoEndereco=%s | estoque_minimo=%s",
        incluir_path,
        saldo_atual_path,
        saldo_endereco_path,
        estoque_minimo_path or "NAO_ENCONTRADA",
    )

    id_token = service_token
    safe_email = _safe_text(auth_email or email) or "automus-sql-only"
    banco = {}
    hash_banco_antes = _json_hash(banco)
    log.info("AUTOMUS_SQL_ONLY: pulando autenticacao e leitura do banco legado.")

    dados_anteriores = banco.get("dados") if isinstance(banco.get("dados"), list) else []
    dados_mortos = banco.get("dadosMortos") if isinstance(banco.get("dadosMortos"), list) else []
    ajustes_itens = banco.get("ajustesItens") if isinstance(banco.get("ajustesItens"), dict) else {}
    historico_saldo = banco.get("historicoSaldo") if isinstance(banco.get("historicoSaldo"), dict) else {}
    movimentacoes_mata185_anteriores = banco.get("movimentacoesMata185") if isinstance(banco.get("movimentacoesMata185"), dict) else {}

    incluir = _read_sheet(incluir_path)
    saldo_atual = _read_sheet(saldo_atual_path)
    saldo_endereco = _read_sheet(saldo_endereco_path)
    estoque_minimo = _read_sheet(estoque_minimo_path) if estoque_minimo_path else None
    dados_mortos_planilha = _read_sheet(dados_mortos_planilha_path) if dados_mortos_planilha_path else None
    mata110 = _read_sheet(mata110_path) if mata110_path else None
    mata111 = _read_sheet(mata111_path) if mata111_path else None
    mata112 = _read_sheet(mata112_path, "Saldos a Endere") if mata112_path else None
    mata185 = _read_sheet(mata185_path) if mata185_path else None
    mata185_encerradas = None
    if mata185_path:
        try:
            mata185_encerradas = _read_sheet(mata185_path, "REQUISICOES ENCERRADAS")
        except RuntimeError:
            mata185_encerradas = []
    log.info(
        "TEST_PLANILHAS_LIDAS_OK | incluir_linhas=%s | saldoAtual_linhas=%s | saldoEndereco_linhas=%s | estoque_minimo_linhas=%s",
        len(incluir),
        len(saldo_atual),
        len(saldo_endereco),
        len(estoque_minimo) if estoque_minimo else 0,
    )
    if len(incluir) < 3 or len(saldo_atual) < 2 or len(saldo_endereco) < 2:
        raise RuntimeError(
            "NAO CONFORME: uma ou mais planilhas estao vazias/incompletas para atualizacao."
        )

    novos_dados = _build_items(
        incluir,
        [],
        saldo_atual,
        saldo_endereco,
        estoque_minimo,
        dados_mortos_planilha,
        dados_anteriores + dados_mortos,
        ajustes_itens,
    )
    if not novos_dados:
        raise RuntimeError("NAO CONFORME: geracao de itens resultou em 0 registros.")
    novos_dados, dados_mortos = _separar_dados_mortos_por_descricao(novos_dados, dados_mortos)
    log.info(
        "TEST_GERACAO_ITENS_OK | itensAtivos=%s | dadosMortos=%s",
        len(novos_dados),
        len(dados_mortos),
    )

    agora_ms = int(time.time() * 1000)
    historico_saldo = _atualizar_historico_saldo(
        novos_dados,
        dados_anteriores,
        historico_saldo,
        agora_ms,
    )
    pedidos_compra = _montar_pedidos_compra(mata110, mata111, mata112, novos_dados, historico_saldo)
    _validar_entradas_mata112_refletidas(mata112, novos_dados, pedidos_compra, log)
    movimentacoes_mata185 = _montar_movimentacoes_mata185(
        mata185,
        mata185_encerradas,
        novos_dados,
        movimentacoes_mata185_anteriores,
        agora_ms,
    )
    _aplicar_sugestoes_estoque(novos_dados, historico_saldo, pedidos_compra)

    payload = {
        "dados": novos_dados,
        "dadosMortos": dados_mortos,
        "ajustesItens": ajustes_itens,
        "historicoSaldo": historico_saldo,
        "pedidosCompra": pedidos_compra,
        "movimentacoesMata185": movimentacoes_mata185,
        "ultimaAtualizacao": agora_ms,
        "atualizadoPor": updated_by,
        "atualizacaoAutomatica": True,
        "ultimaAtualizacaoAutomatica": agora_ms,
        "mapeamentoArquivos": {
            "mata105": "incluir.xlsx",
            "mata225": "saldo.atual.xlsx",
            "mata226": "saldo.por.endereco.xlsx",
            **({"dados_mortos": "dados.mortos.xlsx"} if dados_mortos_planilha_path else {}),
            **({"estoque_minimo": "estoque_minimo.xlsx"} if estoque_minimo_path else {}),
            **({"mata110": "mata110.xlsx"} if mata110_path else {}),
            **({"mata111": "mata111.xlsx"} if mata111_path else {}),
            **({"mata112": "mata112.xlsx"} if mata112_path else {}),
            **({"mata185": "mata185.xlsx"} if mata185_path else {}),
        },
        "automus": {
            "executadoEm": datetime.now().isoformat(timespec="seconds"),
            "usuario": safe_email,
        },
    }

    backup_payload = {
        "salvoEm": agora_ms,
        "origem": "automus",
        "hashAntes": hash_banco_antes,
        "dados": banco,
    }
    backup_remote_payload = {
        "salvoEm": agora_ms,
        "origem": "automus",
        "hashAntes": hash_banco_antes,
        "backupLocal": str(backup_local_path) if "backup_local_path" in locals() else "",
        "dadosQuantidade": len(dados_anteriores),
        "dadosMortosQuantidade": len(dados_mortos),
        "tamanhoBackupLocalBytes": 0,
    }

    backup_local_path = _write_local_backup(project_root, backup_payload, agora_ms, log)
    backup_remote_payload["backupLocal"] = str(backup_local_path)
    backup_remote_payload["tamanhoBackupLocalBytes"] = backup_local_path.stat().st_size

    backup_remoto_ok = False

    log.info(
        "BACKUP_GARANTIA_OK | local=%s | remoto=%s",
        backup_local_path,
        "OK" if backup_remoto_ok else "NAO",
    )

    payload = _sanitize_json_value(payload)
    _json_payload(payload)

    sql_update_ok = False
    try:
        sql_result = _post_inventory_to_sql_api(payload, id_token, log)
        if not sql_result:
            raise RuntimeError("API SQL nao confirmou a carga Automus.")
        sql_update_ok = True
    except Exception as exc:
        log.warning("AUTOMUS_SQL_UPDATE_FALHOU | motivo=%s", exc)

    if not sql_update_ok:
        raise RuntimeError("AUTOMUS_SQL_ONLY ativo, mas a escrita SQL falhou.")
    log.info("AUTOMUS_SQL_ONLY_OK | banco legado nao foi atualizado.")
    _mirror_mata112_to_dashboards(mata112_path, project_root, log)
    return


def _setup_cli_logger() -> logging.Logger:
    logger = logging.getLogger("automus_update_cli")
    if logger.handlers:
        return logger
    logger.setLevel(logging.INFO)
    h = logging.StreamHandler()
    h.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
    logger.addHandler(h)
    return logger


def main() -> None:
    parser = argparse.ArgumentParser(description="Automus: atualiza o estoque no SQL sem usar navegador.")
    parser.add_argument(
        "--config",
        default=str(Path(__file__).resolve().parent / "automus_config.json"),
        help="Caminho do JSON com email/password.",
    )
    parser.add_argument(
        "--project-root",
        default=str(Path(__file__).resolve().parent.parent),
        help="Raiz do projeto (onde esta o index.html e pasta downloads).",
    )
    args = parser.parse_args()

    log = _setup_cli_logger()
    run_automus_update(Path(args.config), Path(args.project_root), logger=log)


if __name__ == "__main__":
    main()
