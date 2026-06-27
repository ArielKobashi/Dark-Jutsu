import csv
import logging
import time
import ctypes
import ctypes.wintypes
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ENCERRADAS_SHEET = "REQUISICOES ENCERRADAS"


def _normalize_digits(value: Any, size: int) -> str:
    digits = "".join(ch for ch in str(value or "") if ch.isdigit())
    return digits.zfill(size) if digits else ""


def _key(sa: Any, item: Any) -> str:
    sa_code = _normalize_digits(sa, 6)
    item_code = _normalize_digits(item, 2)
    return f"{sa_code};{item_code}" if sa_code and item_code else ""


def _candidate_mata185_paths(project_root: Path) -> list[Path]:
    return [
        project_root / "downloads" / "mata185.xlsx",
        project_root / "mata185.xlsx",
        Path.home() / "Desktop" / "mata185.xlsx",
        Path.home() / "Downloads" / "mata185.xlsx",
    ]


def _find_mata185(project_root: Path) -> Path:
    candidates = [path for path in _candidate_mata185_paths(project_root) if path.exists()]
    if not candidates:
        raise FileNotFoundError("mata185.xlsx nao encontrada para registrar requisicoes encerradas.")
    return max(candidates, key=lambda path: path.stat().st_mtime)


def _read_keys(path: Path) -> set[str]:
    if not path.exists():
        return set()
    keys: set[str] = set()
    for line in path.read_text(encoding="utf-8-sig", errors="replace").splitlines():
        text = line.strip()
        if not text or "Nr.S" in text:
            continue
        if ";" in text:
            parts = next(csv.reader([text], delimiter=";"))
            if len(parts) >= 2:
                key = _key(parts[0], parts[1])
                if key:
                    keys.add(key)
        else:
            sa = _normalize_digits(text, 6)
            if sa:
                keys.add(f"{sa};")
    return keys


def _write_encerradas_sheet(mata185_path: Path, azul_keys: set[str]) -> dict[str, int]:
    wb = load_workbook(mata185_path)
    try:
        source = wb[wb.sheetnames[0]]
        if ENCERRADAS_SHEET in wb.sheetnames:
            del wb[ENCERRADAS_SHEET]
        target = wb.create_sheet(ENCERRADAS_SHEET)

        max_col = source.max_column
        for col in range(1, max_col + 1):
            target.cell(row=1, column=col).value = source.cell(row=1, column=col).value
        target.cell(row=1, column=max_col + 1).value = "Status Encerramento"
        target.cell(row=1, column=max_col + 2).value = "Chave Azul"

        matched = 0
        rows_seen = 0
        target_row = 2
        for row in range(2, source.max_row + 1):
            rows_seen += 1
            chave = _key(source.cell(row=row, column=1).value, source.cell(row=row, column=2).value)
            if not chave:
                continue
            matches = chave in azul_keys or f"{chave.split(';', 1)[0]};" in azul_keys
            if not matches:
                continue
            for col in range(1, max_col + 1):
                target.cell(row=target_row, column=col).value = source.cell(row=row, column=col).value
            target.cell(row=target_row, column=max_col + 1).value = "Requisicao encerrada"
            target.cell(row=target_row, column=max_col + 2).value = chave
            matched += 1
            target_row += 1

        wb.save(mata185_path)
        return {
            "azuis": len(azul_keys),
            "linhas_mata185": rows_seen,
            "encerradas": matched,
        }
    finally:
        wb.close()


def _enum_windows() -> list[int]:
    user32 = ctypes.windll.user32
    handles: list[int] = []

    @ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_void_p, ctypes.c_void_p)
    def callback(hwnd, _lparam):
        if user32.IsWindowVisible(hwnd):
            length = user32.GetWindowTextLengthW(hwnd)
            if length > 0:
                handles.append(int(hwnd))
        return True

    user32.EnumWindows(callback, 0)
    return handles


def _window_title(hwnd: int) -> str:
    user32 = ctypes.windll.user32
    length = user32.GetWindowTextLengthW(hwnd)
    buffer = ctypes.create_unicode_buffer(length + 1)
    user32.GetWindowTextW(hwnd, buffer, length + 1)
    return buffer.value


def _find_totvs_window() -> tuple[int, tuple[int, int, int, int]]:
    best_hwnd = 0
    best_score = -1
    for hwnd in _enum_windows():
        title = _window_title(hwnd).lower()
        if "totvs" not in title:
            continue
        score = 1
        if "smartclient" in title:
            score += 4
        if "manufatura" in title:
            score += 2
        if "html" in title:
            score += 1
        if score > best_score:
            best_score = score
            best_hwnd = hwnd

    if not best_hwnd:
        raise RuntimeError("Janela TOTVS nao encontrada para verificacao visual dos azuis.")

    rect = ctypes.wintypes.RECT()
    ctypes.windll.user32.GetWindowRect(best_hwnd, ctypes.byref(rect))
    return best_hwnd, (rect.left, rect.top, rect.right, rect.bottom)


def _is_status_blue(rgb: tuple[int, int, int]) -> bool:
    r, g, b = rgb
    return b >= 135 and g >= 90 and b >= r + 35 and b >= g + 15 and r <= 150


def _is_status_green(rgb: tuple[int, int, int]) -> bool:
    r, g, b = rgb
    return g >= 135 and g >= r + 35 and g >= b + 35 and r <= 160 and b <= 170


def _cluster_rows(row_hits: dict[int, dict[str, int]]) -> list[dict[str, int]]:
    clusters: list[list[tuple[int, dict[str, int]]]] = []
    for y in sorted(row_hits):
        if not clusters or y - clusters[-1][-1][0] > 3:
            clusters.append([])
        clusters[-1].append((y, row_hits[y]))

    rows = []
    for cluster in clusters:
        blue = sum(item["blue"] for _, item in cluster)
        green = sum(item["green"] for _, item in cluster)
        total = blue + green
        if total < 8:
            continue
        weighted_y = sum(y * (item["blue"] + item["green"]) for y, item in cluster) / total
        rows.append({"y": int(round(weighted_y)), "blue": blue, "green": green})
    return rows


def _read_mata185_keys(mata185_path: Path) -> list[str]:
    wb = load_workbook(mata185_path, read_only=True, data_only=True)
    try:
        sheet = wb[wb.sheetnames[0]]
        keys = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            key = _key(row[0] if len(row) > 0 else None, row[1] if len(row) > 1 else None)
            if key:
                keys.append(key)
        return keys
    finally:
        wb.close()


def _ocr_key_from_row(image, row_y: int) -> str:
    try:
        import pytesseract  # type: ignore
    except Exception:
        return ""

    try:
        crop = image.crop((55, max(0, row_y - 13), 230, min(image.size[1], row_y + 14)))
        text = pytesseract.image_to_string(crop, lang="por+eng", config="--psm 7")
    except Exception:
        return ""

    numbers = ["".join(ch for ch in part if ch.isdigit()) for part in text.replace(".", " ").split()]
    numbers = [item for item in numbers if item]
    if len(numbers) < 2:
        return ""
    return _key(numbers[0], numbers[1])


def _capture_totvs_image(rect):
    try:
        from PIL import ImageGrab  # type: ignore
    except Exception as exc:
        raise RuntimeError("Pillow/ImageGrab indisponivel para verificacao visual dos azuis.") from exc

    return ImageGrab.grab(bbox=rect).convert("RGB")


def _status_rows_from_image(image) -> list[dict[str, int]]:
    width, height = image.size
    pixels = image.load()
    row_hits: dict[int, dict[str, int]] = {}
    x_min = 6
    x_max = min(42, width)
    y_min = min(max(160, int(height * 0.20)), height)
    y_max = max(y_min, min(height - 35, int(height * 0.95)))

    for y in range(y_min, y_max):
        blue = 0
        green = 0
        for x in range(x_min, x_max):
            rgb = pixels[x, y]
            if _is_status_blue(rgb):
                blue += 1
            elif _is_status_green(rgb):
                green += 1
        if blue + green >= 3:
            row_hits[y] = {"blue": blue, "green": green}

    return _cluster_rows(row_hits)


def _image_fingerprint(image) -> str:
    width, height = image.size
    pixels = image.load()
    parts = []
    x_start = 0
    x_end = min(width, 260)
    y_start = min(max(170, int(height * 0.22)), height)
    y_end = max(y_start, min(height - 40, int(height * 0.92)))
    step_x = max(1, (x_end - x_start) // 18)
    step_y = max(1, (y_end - y_start) // 28)
    for y in range(y_start, y_end, step_y):
        for x in range(x_start, x_end, step_x):
            r, g, b = pixels[x, y]
            parts.append(f"{r//16:x}{g//16:x}{b//16:x}")
    return "".join(parts)


def _send_page_down(rect):
    left, top, right, bottom = rect
    x = int(left + (right - left) * 0.50)
    y = int(top + (bottom - top) * 0.55)
    user32 = ctypes.windll.user32
    user32.SetCursorPos(x, y)
    time.sleep(0.03)
    mouseeventf_leftdown = 0x0002
    mouseeventf_leftup = 0x0004
    mouseeventf_wheel = 0x0800
    user32.mouse_event(mouseeventf_leftdown, 0, 0, 0, 0)
    time.sleep(0.02)
    user32.mouse_event(mouseeventf_leftup, 0, 0, 0, 0)
    time.sleep(0.05)
    for _ in range(8):
        user32.mouse_event(mouseeventf_wheel, 0, 0, -120, 0)
        time.sleep(0.015)


def _send_home(rect):
    left, top, right, bottom = rect
    x = int(left + (right - left) * 0.50)
    y = int(top + (bottom - top) * 0.55)
    user32 = ctypes.windll.user32
    user32.SetCursorPos(x, y)
    time.sleep(0.03)
    vk_home = 0x24
    vk_control = 0x11
    user32.keybd_event(vk_control, 0, 0, 0)
    user32.keybd_event(vk_home, 0, 0, 0)
    time.sleep(0.02)
    user32.keybd_event(vk_home, 0, 0x0002, 0)
    user32.keybd_event(vk_control, 0, 0x0002, 0)


def _visual_scan_azuis(
    logger: logging.Logger,
    output_dir: Path,
    mata185_path: Path,
    max_pages: int = 1300,
    stable_limit: int = 8,
    page_delay: float = 0.55,
) -> set[str]:
    hwnd, rect = _find_totvs_window()
    logger.info("===== AZUL VISUAL 01/08 | janela TOTVS encontrada | hwnd=%s | rect=%s =====", hwnd, rect)
    try:
        ctypes.windll.user32.SetForegroundWindow(hwnd)
    except Exception:
        pass
    time.sleep(0.4)
    logger.info("===== AZUL VISUAL 02/08 | mantendo posicao atual da grade, sem Ctrl+Home =====")
    time.sleep(0.3)

    mata_keys = _read_mata185_keys(mata185_path)
    logger.info(
        "===== AZUL VISUAL 03/08 | mata185 carregada | linhas_estimadas=%s | arquivo=%s =====",
        len(mata_keys),
        mata185_path,
    )
    azul_keys: set[str] = set()
    debug_lines = ["page;visible_row;estimated_mata185_row;y;blue_pixels;green_pixels;key;source;fingerprint"]
    seen_fingerprints: dict[str, int] = {}
    stable_count = 0
    estimated_first_index = 0
    last_visible_count = 0
    pages_scanned = 0

    logger.info("===== AZUL VISUAL 04/08 | iniciando varredura por paginas | max_pages=%s =====", max_pages)
    for page in range(1, max_pages + 1):
        pages_scanned = page
        image = _capture_totvs_image(rect)
        if page == 1:
            try:
                image.save(output_dir / "verificador-azul-visual-inicio.png")
                logger.info(
                    "===== AZUL VISUAL DEBUG | primeira captura salva em %s =====",
                    output_dir / "verificador-azul-visual-inicio.png",
                )
            except Exception:
                pass

        fingerprint = _image_fingerprint(image)
        if fingerprint in seen_fingerprints:
            stable_count += 1
        else:
            seen_fingerprints[fingerprint] = page
            stable_count = 0

        rows = _status_rows_from_image(image)
        if not rows:
            logger.warning("AZUL: pagina %s sem indicadores visuais detectados.", page)
        else:
            last_visible_count = len(rows)

        blue_rows_this_page = 0
        for index, row in enumerate(rows):
            estimated_index = estimated_first_index + index
            if row["blue"] <= row["green"]:
                continue
            blue_rows_this_page += 1

            key = _ocr_key_from_row(image, row["y"])
            source = "ocr"
            if not key:
                key = mata_keys[estimated_index] if estimated_index < len(mata_keys) else ""
                source = "mata185_ordem_visual"
            if key:
                azul_keys.add(key)
                debug_lines.append(
                    f"{page};{index + 1};{estimated_index + 1};{row['y']};"
                    f"{row['blue']};{row['green']};{key};{source};{fingerprint[:24]}"
                )

        if page <= 5 or page % 10 == 0 or blue_rows_this_page:
            logger.info(
                "===== AZUL VISUAL PAGINA %s | linhas_visiveis=%s | azuis_na_tela=%s | azuis_total=%s | linha_estimada=%s/%s | repeticoes=%s =====",
                page,
                len(rows),
                blue_rows_this_page,
                len(azul_keys),
                min(estimated_first_index + max(last_visible_count, 0), len(mata_keys)),
                len(mata_keys),
                stable_count,
            )

        if stable_count >= stable_limit:
            logger.info("===== AZUL VISUAL 05/08 | fim da lista detectado por tela repetida | repeticoes=%s =====", stable_count)
            break
        if estimated_first_index >= len(mata_keys):
            logger.info("===== AZUL VISUAL 05/08 | fim estimado da mata185 atingido =====")
            break

        step_rows = max(1, last_visible_count - 1)
        estimated_first_index += step_rows
        if page <= 5 or page % 10 == 0:
            logger.info(
                "===== AZUL VISUAL ROLAGEM | mouse wheel na grade | avanco_estimado=%s linhas | proxima_linha=%s =====",
                step_rows,
                estimated_first_index + 1,
            )
        _send_page_down(rect)
        time.sleep(page_delay)

    try:
        final_image = _capture_totvs_image(rect)
        final_image.save(output_dir / "verificador-azul-visual-fim.png")
    except Exception:
        pass

    (output_dir / "lista-azuis.txt").write_text(
        "\n".join(sorted(azul_keys)) + ("\n" if azul_keys else ""),
        encoding="utf-8",
    )
    logger.info("===== AZUL VISUAL 06/08 | lista salva | %s =====", output_dir / "lista-azuis.txt")
    (output_dir / "lista-azuis-visual-debug.csv").write_text(
        "\n".join(debug_lines) + "\n",
        encoding="utf-8",
    )
    logger.info("===== AZUL VISUAL 07/08 | debug salvo | %s =====", output_dir / "lista-azuis-visual-debug.csv")
    logger.info(
        "===== AZUL VISUAL 08/08 | concluida | paginas=%s | azuis=%s | debug=%s =====",
        pages_scanned,
        len(azul_keys),
        output_dir / "lista-azuis-visual-debug.csv",
    )
    return azul_keys


def registrar_requisicoes_encerradas(
    logger: logging.Logger,
    base: Path,
    project_root: Path,
    timeout_seconds: int = 900,
) -> dict[str, int]:
    output_dir = project_root / "downloads"
    output_dir.mkdir(parents=True, exist_ok=True)

    logger.info("===== AZUL VISUAL START | iniciando verificador visual de requisicoes encerradas =====")
    started = time.time()
    mata185_path = _find_mata185(project_root)
    keys: set[str] = set()
    try:
        keys = _visual_scan_azuis(logger, output_dir, mata185_path)
        logger.info("===== AZUL VISUAL DONE | verificador visual concluido em %.1fs =====", time.time() - started)
    except Exception as exc:
        logger.warning(
            "===== AZUL VISUAL ERRO | falhou: %s ===== "
            "A aba de encerradas sera recriada vazia e o envio ao Firebase continuara.",
            exc,
        )

    result = _write_encerradas_sheet(mata185_path, keys)
    logger.info(
        "===== AZUL VISUAL PLANILHA | aba %s atualizada em %s | azuis=%s | encerradas=%s =====",
        ENCERRADAS_SHEET,
        mata185_path,
        result["azuis"],
        result["encerradas"],
    )
    return result
